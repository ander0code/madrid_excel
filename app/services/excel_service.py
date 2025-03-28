import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import sys
import io
import os
from typing import List, Dict, Any, Optional

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.formatters import formatear_dias_teletrabajo
from config import settings


COLOR_ROJO = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
COLOR_VERDE = PatternFill(start_color="538D22", end_color="538D22", fill_type="solid")
COLOR_ENCABEZADO = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
COLOR_NEGRO = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Color negro para teletrabajo
COLOR_GRIS_CLARO = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # Color gris para NM
COLOR_TELETRABAJO = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Gris claro para teletrabajo
COLOR_AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo para tolerancia

MARGEN_TOLERANCIA = 5  

DIAS_SEMANA_MAP = {
    0: "lun",  # Lunes
    1: "mar",  # Martes
    2: "mie",  # Miércoles
    3: "jue",  # Jueves
    4: "vier", # Viernes
    5: "sab",  # Sábado
    6: "dom"   # Domingo
}

async def generate_excel_report(empleados_data: List[Dict[str, Any]], fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None) -> bytes:
    """
    Genera un archivo Excel con las marcaciones de los empleados y lo devuelve como bytes.
    
    Args:
        empleados_data: Lista de datos de empleados con sus marcaciones
        fecha_inicio: Fecha inicial en formato YYYY-MM-DD
        fecha_fin: Fecha final en formato YYYY-MM-DD
        
    Returns:
        Bytes del archivo Excel generado
    """
    try:
        print(f"Generando Excel con {len(empleados_data)} empleados...")
        print(f"Rango de fechas: {fecha_inicio} a {fecha_fin}")
        print(f"Margen de tolerancia configurado: {MARGEN_TOLERANCIA} minutos")
        
        if fecha_inicio and fecha_fin:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
                fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
                
                if fecha_fin_dt < fecha_inicio_dt:
                    print("Advertencia: Fecha de fin anterior a fecha de inicio. Invirtiendo el rango.")
                    fecha_inicio_dt, fecha_fin_dt = fecha_fin_dt, fecha_inicio_dt
                
                delta_dias = (fecha_fin_dt - fecha_inicio_dt).days + 1
                if delta_dias > 31:
                    print(f"Advertencia: El rango de {delta_dias} días es muy amplio. Limitando a 31 días.")
                    fecha_fin_dt = fecha_inicio_dt + timedelta(days=30) 
                
                print(f"Generando reporte para {(fecha_fin_dt - fecha_inicio_dt).days + 1} días")
                usar_fechas_dinamicas = True
            except ValueError as e:
                print(f"Error al parsear fechas: {str(e)}. Usando fechas por defecto.")
                usar_fechas_dinamicas = False
        else:
            print("No se proporcionaron fechas completas. Usando fechas por defecto.")
            usar_fechas_dinamicas = False
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = settings.EXCEL_SHEET_TITLE

        ws.merge_cells('A1:Z1')
        ws['A1'] = "REPORTE DE CONTROL DE MARCACIONES Y ASISTENCIA DEL PERSONAL"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left')

        subtitulos = [
            "Gestión del Talento Humano",
            "MADRID INGENIEROS SAC",
            "Horario General: 08:30 AM - 6:30 PM",
            "Horario Área de Ventas: 09:00AM - 07:00PM",
            "Horario Practicantes Pre (08:30AM - 03:30PM)"
        ]
        for idx, texto in enumerate(subtitulos, start=2):
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=50)
            ws.cell(row=idx, column=1, value=texto).alignment = Alignment(horizontal='left')

        encabezados = [
            "N.", "DNI", "TRABAJADOR", "FECHA INGRESO", "FECHA DE CESE", "CARGO",
            "AREA", "GERENCIA", "ESTADO", "REGISTRO", "DIAS DE LABORES", "DSO",
            "HORARIO OFICIAL", "DÍAS DE TELETRABAJO JD 2025"
        ]
        col = 1
        for encabezado in encabezados:
            ws.merge_cells(start_row=8, start_column=col, end_row=10, end_column=col)
            ws.cell(row=8, column=col, value=encabezado).alignment = Alignment(horizontal='center', vertical='center')
            col += 1

        fechas_dias = []
        fecha_col_map = {}
        dia_semana_map = {}  # Para mapear fecha ISO con el día de la semana
        # Diccionario para rastrear las columnas que son TAR o EXT
        columnas_tardanza_extension = {}
        
        # Crear mapeo de fechas a días de la semana
        todas_fechas = []
        
        if usar_fechas_dinamicas:
            fecha_actual = fecha_inicio_dt
            while fecha_actual <= fecha_fin_dt:
                dias_semana = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO", "DOMINGO"]
                dia_nombre = dias_semana[fecha_actual.weekday()]
                fecha_iso = fecha_actual.strftime("%Y-%m-%d")
                
                # Guardar el día de la semana de cada fecha
                dia_semana = DIAS_SEMANA_MAP[fecha_actual.weekday()]
                dia_semana_map[fecha_iso] = dia_semana
                todas_fechas.append((fecha_iso, dia_semana))
                
                fecha_mostrar = fecha_actual.strftime("%d %B %Y").upper()
                mes_espanol = {
                    "JANUARY": "ENERO", "FEBRUARY": "FEBRERO", "MARCH": "MARZO",
                    "APRIL": "ABRIL", "MAY": "MAYO", "JUNE": "JUNIO",
                    "JULY": "JULIO", "AUGUST": "AGOSTO", "SEPTEMBER": "SEPTIEMBRE",
                    "OCTOBER": "OCTUBRE", "NOVEMBER": "NOVIEMBRE", "DECEMBER": "DICIEMBRE"
                }
                
                for eng, esp in mes_espanol.items():
                    fecha_mostrar = fecha_mostrar.replace(eng, esp)
                
                fechas_dias.append((fecha_mostrar, dia_nombre))
                fecha_actual += timedelta(days=1)
        else:
            fechas_dias = [
                ("03 FEBRERO 2025", "LUNES"), ("04 FEBRERO 2025", "MARTES"), ("05 FEBRERO 2025", "MIÉRCOLES"),
                ("06 FEBRERO 2025", "JUEVES"), ("07 FEBRERO 2025", "VIERNES"), ("08 FEBRERO 2025", "SÁBADO"),
                ("09 FEBRERO 2025", "DOMINGO"), ("10 FEBRERO 2025", "LUNES"), ("11 FEBRERO 2025", "MARTES"),
                ("12 FEBRERO 2025", "MIÉRCOLES"), ("13 FEBRERO 2025", "JUEVES"), ("14 FEBRERO 2025", "VIERNES")
            ]
            
            # Para fechas estáticas, mapear manualmente
            dias_mapeo = [
                ("2025-02-03", "lun"), ("2025-02-04", "mar"), ("2025-02-05", "mie"),
                ("2025-02-06", "jue"), ("2025-02-07", "vier"), ("2025-02-08", "sab"),
                ("2025-02-09", "dom"), ("2025-02-10", "lun"), ("2025-02-11", "mar"),
                ("2025-02-12", "mie"), ("2025-02-13", "jue"), ("2025-02-14", "vier")
            ]
            todas_fechas = dias_mapeo
            for fecha, dia in dias_mapeo:
                dia_semana_map[fecha] = dia
        
        columnas_fecha = ["ING", "TAR", "SALIDA", "EXT"]
        
        for i, (fecha, dia) in enumerate(fechas_dias):
            if usar_fechas_dinamicas:
                fecha_dt = fecha_inicio_dt + timedelta(days=i)
                fecha_iso = fecha_dt.strftime("%Y-%m-%d")
            else:
                fecha_iso = f"2025-02-{int(fecha[:2]):02d}"
            
            fecha_col_map[fecha_iso] = col

            ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+3)
            ws.cell(row=8, column=col, value=fecha).alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col+3)
            ws.cell(row=9, column=col, value=dia).alignment = Alignment(horizontal='center', vertical='center')
            
            for j, sub in enumerate(columnas_fecha):
                celda = ws.cell(row=10, column=col+j, value=sub)
                celda.alignment = Alignment(horizontal='center', vertical='center')
                if sub in ["TAR", "EXT"]:
                    celda.fill = COLOR_ROJO
                    celda.font = Font(color="FFFFFF", bold=True)
                    # Registrar columnas TAR y EXT para aplicar negrita posteriormente
                    columnas_tardanza_extension[col+j] = sub
            col += 4
        
        # Añadir nuevas columnas de cantidades
        col_cant_tardanzas = col
        col_cant_tolerancias = col + 1
        col_cant_faltas = col + 2
        col_total_tardanza = col + 3
        col_total_ausencia = col + 4
        
        # Añadir encabezados para las columnas de cantidades
        ws.merge_cells(start_row=8, start_column=col_cant_tardanzas, end_row=10, end_column=col_cant_tardanzas)
        celda_cant_tard = ws.cell(row=8, column=col_cant_tardanzas, value="CANT. TARDANZAS")
        celda_cant_tard.alignment = Alignment(horizontal='center', vertical='center')
        celda_cant_tard.fill = COLOR_ENCABEZADO
        celda_cant_tard.font = Font(color="FFFFFF", bold=True)
        
        ws.merge_cells(start_row=8, start_column=col_cant_tolerancias, end_row=10, end_column=col_cant_tolerancias)
        celda_cant_toler = ws.cell(row=8, column=col_cant_tolerancias, value="CANT. TOLERANCIAS")
        celda_cant_toler.alignment = Alignment(horizontal='center', vertical='center')
        celda_cant_toler.fill = COLOR_ENCABEZADO
        celda_cant_toler.font = Font(color="FFFFFF", bold=True)
        
        ws.merge_cells(start_row=8, start_column=col_cant_faltas, end_row=10, end_column=col_cant_faltas)
        celda_cant_faltas = ws.cell(row=8, column=col_cant_faltas, value="CANT. FALTAS")
        celda_cant_faltas.alignment = Alignment(horizontal='center', vertical='center')
        celda_cant_faltas.fill = COLOR_ENCABEZADO
        celda_cant_faltas.font = Font(color="FFFFFF", bold=True)
        
        # Añadir encabezados para las columnas de totales con nombres actualizados
        ws.merge_cells(start_row=8, start_column=col_total_tardanza, end_row=10, end_column=col_total_tardanza)
        celda_total_tardanza = ws.cell(row=8, column=col_total_tardanza, value="TOTAL DE MINUTOS DE TARDANZA")
        celda_total_tardanza.alignment = Alignment(horizontal='center', vertical='center')
        celda_total_tardanza.fill = COLOR_ENCABEZADO
        celda_total_tardanza.font = Font(color="FFFFFF", bold=True)
        
        ws.merge_cells(start_row=8, start_column=col_total_ausencia, end_row=10, end_column=col_total_ausencia)
        celda_total_ausencia = ws.cell(row=8, column=col_total_ausencia, value="TOTAL DE MINUTOS DE AUSENCIA")
        celda_total_ausencia.alignment = Alignment(horizontal='center', vertical='center')
        celda_total_ausencia.fill = COLOR_ENCABEZADO
        celda_total_ausencia.font = Font(color="FFFFFF", bold=True)
        
        col += 5  # Actualizar el contador de columnas después de añadir las nuevas columnas

        # Aplicar estilo a todas las celdas de encabezado
        for row in ws.iter_rows(min_row=8, max_row=10, min_col=1, max_col=col-1):
            for celda in row:
                if not celda.fill.start_color.index == "FF0000":  # Si no es una celda roja (TAR, EXT)
                    celda.fill = COLOR_ENCABEZADO
                    celda.font = Font(color="FFFFFF", bold=True)

        fila_actual = 11
        empleados_validos = [e for e in empleados_data if isinstance(e, dict) and e.get("emp_code")]
        
        for idx, empleado in enumerate(empleados_validos, 1):
            try:
                dias_remoto = empleado.get("dias_remoto", [])
                
                fechas_teletrabajo = set()
                for fecha_iso, dia_semana in todas_fechas:
                    if dia_semana in dias_remoto:
                        fechas_teletrabajo.add(fecha_iso)
                
                ws.cell(row=fila_actual, column=1, value=idx)
                ws.cell(row=fila_actual, column=2, value=empleado.get("emp_code", ""))
                
                first_name = empleado.get("first_name", "") or ""
                last_name = empleado.get("last_name", "") or ""
                nombre_completo = f"{first_name} {last_name}".strip()
                ws.cell(row=fila_actual, column=3, value=nombre_completo if nombre_completo else "-")
                
                if empleado.get("hire_date"):
                    try:
                        fecha_ingreso = datetime.strptime(empleado["hire_date"], "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%d/%m/%Y")
                        ws.cell(row=fila_actual, column=4, value=fecha_ingreso)
                    except (ValueError, TypeError):
                        ws.cell(row=fila_actual, column=4, value="-")
                else:
                    ws.cell(row=fila_actual, column=4, value="-")
                
                fecha_cese = None
                tiene_fecha_cese = False
                fecha_cese_str = "-"

                if empleado.get("fecha_cese"):
                    try:
                        fecha_cese = datetime.strptime(empleado["fecha_cese"], "%Y-%m-%dT%H:%M:%S.%fZ")
                        fecha_cese_str = fecha_cese.strftime("%d/%m/%Y")
                        tiene_fecha_cese = True
                    except (ValueError, TypeError):
                        pass
                        
                ws.cell(row=fila_actual, column=5, value=fecha_cese_str)
                
                ws.cell(row=fila_actual, column=6, value=empleado.get("position_name", "-"))
                dept_name = empleado.get("dept_name", "-")
                ws.cell(row=fila_actual, column=7, value=dept_name)
                ws.cell(row=fila_actual, column=8, value=empleado.get("gerencia", "-"))

                if tiene_fecha_cese:
                    estado = "Cesado"
                elif empleado.get("is_unactive", False):
                    estado = "Inactivo"
                else:
                    estado = "Activo"
                
                ws.cell(row=fila_actual, column=9, value=estado)
                
                ws.cell(row=fila_actual, column=10, value=empleado.get("registro", "-"))
                
                dias_labores = empleado.get("dias_labores", "-")
                if dias_labores == "lun-vier":
                    ws.cell(row=fila_actual, column=11, value="LUNES A VIERNES")
                else:
                    ws.cell(row=fila_actual, column=11, value=dias_labores.upper() if dias_labores else "-")
                
                dias_descanso = empleado.get("dias_descanso", "-")
                if dias_descanso == "sab-dom":
                    ws.cell(row=fila_actual, column=12, value="S Y D")
                else:
                    ws.cell(row=fila_actual, column=12, value=dias_descanso.upper() if dias_descanso else "-")
                
                if empleado.get("hora_ingreso") and empleado.get("hora_salida"):
                    horario = f"{empleado['hora_ingreso']}AM - {empleado['hora_salida']}PM"
                    ws.cell(row=fila_actual, column=13, value=horario)
                else:
                    ws.cell(row=fila_actual, column=13, value="-")

                # Agregar los días de teletrabajo formateados
                ws.cell(row=fila_actual, column=14, value=formatear_dias_teletrabajo(dias_remoto))
                
                # Inicializar un diccionario para rastrear las marcaciones por fecha
                marcaciones_por_fecha = {}
                
                # Almacenar todas las marcaciones por fecha
                if "marcaciones" in empleado and isinstance(empleado["marcaciones"], list):
                    for marcacion in empleado["marcaciones"]:
                        try:
                            if not isinstance(marcacion, dict) or "fecha" not in marcacion:
                                continue
                                
                            fecha_marca = None
                            try:
                                fecha_marca = datetime.strptime(marcacion["fecha"], "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%Y-%m-%d")
                            except ValueError:
                                try:
                                    fecha_marca = datetime.strptime(marcacion["fecha"], "%Y-%m-%d").strftime("%Y-%m-%d")
                                except ValueError:
                                    print(f"Error al parsear fecha: {marcacion['fecha']}")
                                    continue
                                    
                            # Almacenar esta marcación
                            marcaciones_por_fecha[fecha_marca] = marcacion
                            
                        except Exception as e:
                            print(f"Error procesando marcación: {str(e)}")
                            continue
                
                # Contadores para tolerancias y tardanzas
                contador_tardanzas = 0
                contador_tolerancias = 0
                
                # Ahora, para cada fecha en el rango, procesarla adecuadamente
                for fecha_iso in fecha_col_map.keys():
                    col_inicio = fecha_col_map[fecha_iso]
                    
                    # Verificar si esta fecha debe ser teletrabajo
                    es_dia_teletrabajo = fecha_iso in fechas_teletrabajo
                    
                    # Verificar si tenemos marcaciones para esta fecha
                    tiene_marcaciones = fecha_iso in marcaciones_por_fecha
                    
                    # Si es día de teletrabajo, aplicar fondo gris claro a las celdas
                    if es_dia_teletrabajo:
                        # Aplicar fondo de teletrabajo a todas las celdas
                        for j in range(4):  # 4 columnas: ING, TAR, SALIDA, EXT
                            celda = ws.cell(row=fila_actual, column=col_inicio+j)
                            celda.fill = COLOR_TELETRABAJO
                            celda.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Si tenemos marcaciones para esta fecha, mostrarlas
                    if tiene_marcaciones:
                        marcacion = marcaciones_por_fecha[fecha_iso]
                        
                        # Celda de entrada - centrada
                        hora_ingreso = marcacion.get("hora_ingreso")
                        if hora_ingreso is None:
                            # Si hora_ingreso es null, escribir "NM" con fondo gris
                            celda_ingreso = ws.cell(row=fila_actual, column=col_inicio, value="NM")
                            celda_ingreso.fill = COLOR_GRIS_CLARO
                            celda_ingreso.font = Font(bold=True)
                        else:
                            celda_ingreso = ws.cell(row=fila_actual, column=col_inicio, value=hora_ingreso)
                            # Si es día de teletrabajo, mantener el fondo de teletrabajo
                            if es_dia_teletrabajo:
                                celda_ingreso.fill = COLOR_TELETRABAJO
                        
                        celda_ingreso.alignment = Alignment(horizontal='center', vertical='center')

                        diferencia_ingreso = marcacion.get("diferencia_ingreso", 0)
                        try:
                            diferencia_ingreso = int(diferencia_ingreso)
                        except (ValueError, TypeError):
                            diferencia_ingreso = 0
                            
                        # Celda de tardanza - centrada
                        celda_tardanza = ws.cell(row=fila_actual, column=col_inicio+1, 
                                              value=str(diferencia_ingreso))
                        celda_tardanza.alignment = Alignment(horizontal='center', vertical='center')

                        # Nueva lógica de color con tolerancia
                        if diferencia_ingreso > MARGEN_TOLERANCIA:
                            celda_tardanza.fill = COLOR_ROJO
                            celda_tardanza.font = Font(color="FFFFFF", bold=True)
                            contador_tardanzas += 1
                        elif diferencia_ingreso > 0:
                            celda_tardanza.fill = COLOR_AMARILLO
                            celda_tardanza.font = Font(bold=True)
                            contador_tolerancias += 1
                        else:
                            if es_dia_teletrabajo:
                                celda_tardanza.fill = COLOR_TELETRABAJO
                            else:
                                celda_tardanza.fill = COLOR_VERDE
                                celda_tardanza.font = Font(color="FFFFFF", bold=True)

                        # Celda de salida - centrada
                        hora_salida = marcacion.get("hora_salida")
                        if hora_salida is None:
                            # Si hora_salida es null, escribir "NM" con fondo gris
                            celda_salida = ws.cell(row=fila_actual, column=col_inicio+2, value="NM")
                            celda_salida.fill = COLOR_GRIS_CLARO
                            celda_salida.font = Font(bold=True)
                        else:
                            celda_salida = ws.cell(row=fila_actual, column=col_inicio+2, value=hora_salida)
                            # Si es día de teletrabajo, mantener el fondo de teletrabajo
                            if es_dia_teletrabajo:
                                celda_salida.fill = COLOR_TELETRABAJO
                        
                        celda_salida.alignment = Alignment(horizontal='center', vertical='center')

                        diferencia_salida = marcacion.get("diferencia_salida", 0)
                        try:
                            diferencia_salida = int(diferencia_salida)
                        except (ValueError, TypeError):
                            diferencia_salida = 0
                            
                        # Celda de extensión - centrada
                        celda_extension = ws.cell(row=fila_actual, column=col_inicio+3, 
                                               value=str(diferencia_salida))
                        celda_extension.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Aplicar color según extensión, pero respetando si es día de teletrabajo
                        if diferencia_salida < 0:
                            celda_extension.fill = COLOR_ROJO
                            celda_extension.font = Font(color="FFFFFF", bold=True)
                        else:
                            if es_dia_teletrabajo:
                                celda_extension.fill = COLOR_TELETRABAJO
                            else:
                                celda_extension.fill = COLOR_VERDE
                                celda_extension.font = Font(color="FFFFFF", bold=True)

                # Calcular totales a partir de las marcaciones si es necesario
                total_tardanza_calculado = 0
                total_ausencia_calculada = 0

                if "marcaciones" in empleado and isinstance(empleado["marcaciones"], list):
                    for marcacion in empleado["marcaciones"]:
                        try:
                            diferencia_ingreso = marcacion.get("diferencia_ingreso", 0)
                            if isinstance(diferencia_ingreso, (int, float)) and diferencia_ingreso > 0:
                                total_tardanza_calculado += diferencia_ingreso
                                
                            diferencia_salida = marcacion.get("diferencia_salida", 0)
                            if isinstance(diferencia_salida, (int, float)) and diferencia_salida < 0:
                                total_ausencia_calculada += diferencia_salida
                        except Exception as e:
                            print(f"Error al calcular totales de marcación: {str(e)}")

                # Extraer los valores de totales y cantidades del JSON
                total_tardanza = empleado.get("total_minutos_tardanzas")
                total_ausencia = empleado.get("total_minutos_salidas_temprano")
                cant_tardanzas = empleado.get("cantidad_tardanzas", contador_tardanzas)
                cant_tolerancias = empleado.get("cantidad_tolerancias", contador_tolerancias)
                cant_faltas = empleado.get("cantidad_faltas", 0)

                # Si los valores originales son None o 0, usar los calculados
                if total_tardanza is None or total_tardanza == 0:
                    total_tardanza = total_tardanza_calculado
                
                if total_ausencia is None or total_ausencia == 0:
                    total_ausencia = total_ausencia_calculada
                
                # Verificar que sean números y convertirlos si es necesario
                if isinstance(total_tardanza, str):
                    try:
                        total_tardanza = int(total_tardanza)
                    except (ValueError, TypeError):
                        total_tardanza = total_tardanza_calculado
                elif not isinstance(total_tardanza, (int, float)):
                    total_tardanza = total_tardanza_calculado
                
                if isinstance(total_ausencia, str):
                    try:
                        total_ausencia = int(total_ausencia)
                    except (ValueError, TypeError):
                        total_ausencia = total_ausencia_calculada
                elif not isinstance(total_ausencia, (int, float)):
                    total_ausencia = total_ausencia_calculada
                
                # Añadir las nuevas columnas de cantidades
                celda_cant_tard = ws.cell(row=fila_actual, column=col_cant_tardanzas, value=cant_tardanzas)
                celda_cant_tard.alignment = Alignment(horizontal='center', vertical='center')
                
                celda_cant_toler = ws.cell(row=fila_actual, column=col_cant_tolerancias, value=cant_tolerancias)
                celda_cant_toler.alignment = Alignment(horizontal='center', vertical='center')
                
                celda_cant_faltas = ws.cell(row=fila_actual, column=col_cant_faltas, value=cant_faltas)
                celda_cant_faltas.alignment = Alignment(horizontal='center', vertical='center')
                
                # Celda Total Tardanza
                celda_total_tard = ws.cell(row=fila_actual, column=col_total_tardanza, value=total_tardanza)
                celda_total_tard.alignment = Alignment(horizontal='center', vertical='center')
                if total_tardanza > 0:
                    celda_total_tard.fill = COLOR_ROJO
                    celda_total_tard.font = Font(color="FFFFFF", bold=True)
                else:
                    celda_total_tard.fill = COLOR_VERDE
                    celda_total_tard.font = Font(color="FFFFFF", bold=True)
                
                # Celda Total Ausencia
                celda_total_aus = ws.cell(row=fila_actual, column=col_total_ausencia, value=total_ausencia)
                celda_total_aus.alignment = Alignment(horizontal='center', vertical='center')
                if total_ausencia < 0:
                    celda_total_aus.fill = COLOR_ROJO
                    celda_total_aus.font = Font(color="FFFFFF", bold=True)
                else:
                    celda_total_aus.fill = COLOR_VERDE
                    celda_total_aus.font = Font(color="FFFFFF", bold=True)
            
            except Exception as e:
                print(f"Error procesando empleado {idx}: {str(e)}")
                import traceback
                traceback.print_exc()
            
            fila_actual += 1

        # Centrar todas las celdas de las columnas generadas por rango de fechas
        # (a partir de la columna 15)
        for row in ws.iter_rows(min_row=11, max_row=fila_actual-1, min_col=15, max_col=col-1):
            for cell in row:
                if not cell.alignment.horizontal:  # Si no tiene alineación definida
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        for row in ws.iter_rows(min_row=8, max_row=fila_actual-1, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        anchos_personalizados = {
            'A': 5, 'B': 12, 'C': 25, 'D': 15, 'E': 15, 'F': 20,
            'G': 15, 'H': 10, 'I': 12, 'J': 12, 'K': 18, 'L': 10,
            'M': 20, 'N': 20
        }
        
        for letra, ancho in anchos_personalizados.items():
            ws.column_dimensions[letra].width = ancho
            
        # Configurar el ancho de las columnas de fechas
        for idx in range(15, col_cant_tardanzas):  # Columnas de fechas
            ws.column_dimensions[get_column_letter(idx)].width = 10
            
        # Configurar el ancho de las nuevas columnas
        ws.column_dimensions[get_column_letter(col_cant_tardanzas)].width = 15
        ws.column_dimensions[get_column_letter(col_cant_tolerancias)].width = 15
        ws.column_dimensions[get_column_letter(col_cant_faltas)].width = 15
        ws.column_dimensions[get_column_letter(col_total_tardanza)].width = 15
        ws.column_dimensions[get_column_letter(col_total_ausencia)].width = 15

        print("Generando bytes del Excel...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_bytes = output.getvalue()
        
        print(f"Excel generado correctamente. Tamaño: {len(excel_bytes) / 1024:.2f} KB")
        return excel_bytes
            
    except Exception as e:
        print(f"Error al generar el Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e